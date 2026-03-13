import openpyxl
import json
import unicodedata
import re

def normalizar(nome):
    if not nome:
        return ''
    nome = str(nome).strip().lower()
    nome = unicodedata.normalize('NFKD', nome)
    nome = ''.join(c for c in nome if not unicodedata.combining(c))
    nome = re.sub(r'\s+', ' ', nome)
    return nome

# === 1. Carregar Excel (251 matrículas = todos os contratos ativos em fev/26) ===
wb = openpyxl.load_workbook(
    r'c:\Users\hugog\OneDrive\Desktop\Projects\LA Music\LAperformanceReport\data\barra\matricula_barra_fevereiro.xlsx',
    data_only=True
)
ws = wb.active
headers = [cell.value for cell in ws[1]]

# Cada linha do Excel = 1 matrícula/contrato
excel_mats = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None for v in row):
        r = dict(zip(headers, row))
        nome = str(r.get('Nome do Aluno', '') or '').strip()
        if nome:
            excel_mats.append({
                'nome_original': nome,
                'nome_norm': normalizar(nome),
                'curso': str(r.get('Curso', '') or '').strip(),
                'mensalidade': r.get('Mensalidade', 0) or 0,
                'data_interrupcao': r.get('Data de Interrupção', None),
            })

print(f"Total matrículas no Excel: {len(excel_mats)}")

# Nomes únicos no Excel (alunos únicos)
nomes_excel_unicos = {}
for m in excel_mats:
    n = m['nome_norm']
    if n not in nomes_excel_unicos:
        nomes_excel_unicos[n] = []
    nomes_excel_unicos[n].append(m)

print(f"Nomes únicos no Excel: {len(nomes_excel_unicos)}")
alunos_excel_2_mats = {n: mats for n, mats in nomes_excel_unicos.items() if len(mats) >= 2}
print(f"Alunos com 2+ matrículas no Excel: {len(alunos_excel_2_mats)}")

# === 2. Carregar banco (248 registros) ===
with open(r'C:\Users\hugog\.gemini\antigravity\brain\583d30eb-e69c-4a3b-8c76-251fe179fdea\.system_generated\steps\167\output.txt', 'r', encoding='utf-8') as f:
    raw = f.read()

inicio = raw.find('[{')
fim = raw.rfind('}]') + 2
banco_data = json.loads(raw[inicio:fim])

# Nomes únicos no banco (excl. 2º curso)
banco_primario = [a for a in banco_data if not a.get('is_segundo_curso')]
banco_segundo_curso = [a for a in banco_data if a.get('is_segundo_curso')]

nomes_banco_primario = {normalizar(a['nome']): a for a in banco_primario}
nomes_banco_todos = {}
for a in banco_data:
    n = normalizar(a['nome'])
    if n not in nomes_banco_todos:
        nomes_banco_todos[n] = []
    nomes_banco_todos[n].append(a)

print(f"\nTotal banco: {len(banco_data)}")
print(f"Banco primário (sem 2º curso): {len(banco_primario)}")
print(f"Banco 2º curso: {len(banco_segundo_curso)}")
print(f"Nomes únicos banco: {len(nomes_banco_todos)}")

# === 3. Cruzamento ===
print("\n" + "="*60)
print("NO EXCEL MAS NAO NO BANCO (nome_norm ausente no banco)")
print("="*60)
so_excel = []
for nome_norm in nomes_excel_unicos:
    if nome_norm not in nomes_banco_todos:
        so_excel.append((nome_norm, nomes_excel_unicos[nome_norm]))

print(f"Total: {len(so_excel)}")
for nome_norm, mats in sorted(so_excel):
    for m in mats:
        print(f"  [{m['nome_original']}] | {m['curso']} | R${m['mensalidade']}")

print("\n" + "="*60)
print("NO BANCO MAS NAO NO EXCEL (nome_norm ausente no Excel)")
print("="*60)
so_banco = []
for nome_norm in nomes_banco_todos:
    if nome_norm not in nomes_excel_unicos:
        so_banco.append((nome_norm, nomes_banco_todos[nome_norm]))

print(f"Total: {len(so_banco)}")
for nome_norm, registros in sorted(so_banco):
    for a in registros:
        segundo = " [2ºCURSO]" if a.get('is_segundo_curso') else ""
        print(f"  [{a['nome']}]{segundo} | status:{a['status']} | R${a.get('valor_parcela',0)} | matric:{a.get('data_matricula','')}")

print("\n" + "="*60)
print("ALUNOS COM 2+ MATRÍCULAS NO EXCEL vs BANCO")
print("="*60)
print(f"Excel: {len(alunos_excel_2_mats)} alunos com 2+ matrículas")
for nome, mats in sorted(alunos_excel_2_mats.items()):
    cursos = [m['curso'] for m in mats]
    banco_regs = nomes_banco_todos.get(nome, [])
    banco_count = len(banco_regs)
    print(f"  {mats[0]['nome_original']}: {len(mats)} no Excel | {banco_count} no banco ({cursos})")

print("\n" + "="*60)
print("RESUMO FINAL")
print("="*60)
print(f"Excel: {len(excel_mats)} matrículas / {len(nomes_excel_unicos)} alunos únicos")
print(f"Banco: {len(banco_data)} registros / {len(nomes_banco_todos)} alunos únicos")
print(f"Só no Excel: {len(so_excel)} alunos")
print(f"Só no Banco: {len(so_banco)} alunos")
